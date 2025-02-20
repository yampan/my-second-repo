import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill

wb = openpyxl.Workbook() # workbookの作成

# 罫線の設定
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)

# 背景色設定
color = PatternFill(patternType='solid', fgColor='D9E1F2')

# 1つ目のシート作成
test1_data = [
    ["ID", "NAME"],
    [1,"AAA"],
    [2,"BBB"],
    [3,"CCC"],
    [4,"DDD"],
    ]
ws = wb["Sheet"] # workbook作成時に「Sheet」シートが作成されているので、それを参照
ws.title = "test1" # sheet名変更
for row_i, row_data in enumerate(test1_data): # 行ごとのデータ取得
    for col_i, data in enumerate(row_data): #行内の列ごとのデータ取得
        col_letter = chr(ord("A")+col_i) # 列アルファベット名の生成(A,B,...とAにcol_iを足して自動生成)
        address = f"{col_letter}{row_i+1}" # セルの位置
        ws[address] = data # 値格納
        ws[address].border = border # 罫線の設定
        if row_i<1:
            ws[address].fill = color # 背景色の設定

# 2つ目のシート作成
color = PatternFill(patternType='solid', fgColor='FCE4D6') # 背景色変更
test2_data = [
    ["ID", "NAME", "", "Address"],
    ["", "姓", "名", ""],
    [1,"A", "AA", "Tokyo"],
    [2,"B", "BB", "Tokyo"],
    [3,"C", "CC", "Osaka"],
    [4,"D", "DD", "Aichi"],
    ]
ws = wb.create_sheet(title="test2") # Sheetの作成
num_empty_row = 1 # 空の行をいくつ挿入するか
num_empty_col = 1 # 空の列をいくつ挿入するか
for row_i, row_data in enumerate(test2_data): # 行ごとのデータ取得
    for col_i, data in enumerate(row_data): #行内の列ごとのデータ取得
        col_letter = chr(ord("A")+col_i+num_empty_col) # 列アルファベット名の生成(A,B,...とAにcol_iとnum_empty_colを足して自動生成)
        address = f"{col_letter}{row_i+1+num_empty_row}" # セルの位置
        ws[address] = data # 値格納
        ws[address].border = border # 罫線の設定
        if row_i<2:
            ws[address].fill = color # 背景色の設定
# セル結合
ws.merge_cells("B2:B3")
ws.merge_cells("C2:D2")
ws.merge_cells("E2:E3")

wb.save("test.xlsx")
