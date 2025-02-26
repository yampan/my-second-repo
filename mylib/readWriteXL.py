### readWriteXl.py
'''
セルの座標 (x, y) は 1 始まりで指定します。
openpyxl で読み書きできるのは .xlsx 形式のファイルのみです。
'''
# module
from openpyxl import load_workbook, styles
import pytz, datetime
import openpyxl
import sys

# j_map: {'var_name' : (81:'Dose') }
j_map = {"id":(112,'ID'), 'kannri_id':(1,'院内管理コード') , 'name':(114, '名前'), 
    'sex':(2,'性別'),'disease':(10, '疾患名'),'dis_icdo':(13, '原発部位ICD-Oコード'),
    'pathology':(14, '病理組織'), 'path_icdo':(15,'病理組織ICD-Oコード'),
    'st_date':(43, '外部照射開始日'), 'en_date':(44, '外部照射終了日'), 
    'dose':(45,'外部照射総線量'),'days':(46,'外部照射日数'),'frac':(47,'外部照射分割回数'),
    'perday':(48, '一日あたり照射回数'),'comp':(85,'放射線治療完遂度'),
    'status':(87,'生死の状況'), 'final_d': (88,'最終確認日') }

# log sheet header
log_header = ["ROW", "COLUMN", "Original value", "New value", "Item name", "Datetime" ]
log_width =  [6, 9, 20, 20, 15, 25]



def xlwidth(ws, col:int, w:int):
    """
    ワークシートwsのcol の幅を設定する。
    Args:
        ws: worksheet
        col: 対象とする column
        w: 幅
    return:
        w: 幅
    """
    column_letter = openpyxl.utils.get_column_letter(col)
    #print(f"col:{col} => letter:{column_letter}")
    ws.column_dimensions[column_letter].width = w # 単位：文字数
    return w


def xlcolor(ws, row:int, col:int, color="ACEBF0"):
    """
    ワークシートwsの(row, col) の色を設定する。
    Args:
        ws: worksheet
        row, col: 対象とする celll(row, col)
        color: 設定する色
    return:
        color: 色
    """
    fill_color1 = styles.PatternFill(fgColor=color, fill_type="solid") 
    cell = ws.cell(row=row, column=col)
    cell.fill = fill_color1
    return color

def trans2(ws, row, d:dict, ws_log, deb=0):
    # d = {"name": value} => col, val
    name = None
    for name, val in d.items():
        pass
    if name is None:
        print(f"trans2: #27 name None ERROR")
        return
    # ---
    print(f"trans2: #30 name = {name}, val = {val}")
    disp = None
    for k in j_map.keys():
        if name == k:
            (col, disp) = j_map[k]
    
    if disp is None:
        print(f"trans2: #35 disp None ERROR")
        return    
    # === goto trans()
    if deb: print(f"trans2: #40 row={row}, col={col}, val={val}, disp={disp}")
    return trans(ws, row, col, val, ws_log, disp)
    
def trans(ws, row, col, val, ws_log, disp="null"):
    """
    ワークシートwsの(row, col)の値を(val)に変更して、ws_logにlogを記録する。（append)
    Args:
        ws: worksheet
        row, val: 対象とするCELL
        val: 書き込む値
        ws_log: 記録するworksheet
        disp: 項目名(default: null)
    return:
        log: ws_logに書き込んだ値。
    """
    cell = ws.cell(row=row, column=col)
    print(f"trans: #56 cell = {cell}")
    old_v = cell.value
    cell.value = val
    
    # --- trans
    log_dat = [row, col, old_v, val, disp, JST()]
    ws_log.append(log_dat)
    return log_dat


# JST (日本標準時) のタイムゾーンを取得
def JST():
    jst = pytz.timezone('Asia/Tokyo')
    now = datetime.datetime.now(jst) # 現在の時刻をJSTで取得
    #now = now.strftime('%Y-%m-%d %H:%M:%S %Z%z') # 表示形式をカスタマイズ
    return now.strftime('%Y-%m-%d %H:%M:%S (%Z)') # 表示形式をカスタマイズ    
    
def JSTfn():
    """
    JST-string suitable for filename 
    """
    jst = pytz.timezone('Asia/Tokyo')
    now = datetime.datetime.now(jst) # 現在の時刻をJSTで取得
    
    return now.strftime('%m-%d_%H%M') # 表示形式をカスタマイズ    
    

def copy_worksheet_with_styles(input_file="mylib\\aaa.xlsx", 
                               output_file="mylib\\bbb.xlsx"):
    """
    ワークシートをコピーし、スタイルも保持する関数

    Args:
        input_file (str): 入力Excelファイル名
        output_file (str): 出力Excelファイル名
    """
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb["test1"]

        # ワークシートをコピー
        ws2 = wb.copy_worksheet(ws)
        ws2.title = "test2" #コピー後にシート名を変更

        wb.save(output_file)
        print("ワークシートをコピーし、スタイルも保持しました。")
        return wb, ws, ws2
    
    except FileNotFoundError:
        print(f"ファイル '{input_file}' が見つかりません。")
    except KeyError:
        print("指定されたシートが見つかりません。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")

### ソート
def sort2(wb, ws, key, deb=0):
    """
    sort keyは1 or 2個。row dataの最後にindexを付加
    args:
        wb: workbook
        ws: worksheet
        key:int or list -  key no (1始まり)
    return:
        ws: sorted ws
    """
    # get all cell
    header = []
    data = []
    index_col = ws.max_column+1
    for n, row in enumerate(ws, start=1): # 各行のループ
        cell = ws.cell(row=n, column=index_col)
        if n == 1: 
            header = row
            cell.value = 'index'
            header = header + (cell,)
        else:
            cell.value=n
            row = row + (cell,)
            data.append(row)
    # ---
    if deb:
        print(f"HDR {header}")
        for n, i in enumerate(data, start=2):
            print(f"{n}): {i}")
    
    # sort
    if type(key) is int:
        data.sort(key=lambda row: (row[key-1].value)) 
    else:
        data.sort(key=lambda row: (row[key[0]-1].value,  row[key[1]-1].value)) 
    print(f"#252 key = {key}")
    for n, i in enumerate(data, start=2):
        print(f"{n}): {i}")
           
    #ws2['A1'] = row2list(header)[0] # OK
    #ws2[1] = row2list(header) # NG
    #print(f"#260 'A1:A2' = {ws2['A1:B1']}") # Cell get OK, cell write NG
    
    # 各セルから値を抽出して、　data_v に保存する。
    data_v = []
    a = []
    for i in header:
        a.append(i.value) 
    data_v.append(a)
    for row in data:
        a = []
        for col in row:
            a.append(col.value)
        data_v.append(a)
        
    # data_v から値をもどす。
    for y, row in enumerate(data_v, start=1):
        for x, col in enumerate(row, start=1):
            ws.cell(row=y, column=x).value = col
    
    # sort column に色を付ける
    fill_color1 = styles.PatternFill(fgColor="ACEBF0", fill_type="solid") # 青色単色
    fill_color2 = styles.PatternFill(fgColor="FFFF66", fill_type="solid") # 黄色
    color = [fill_color1, fill_color2]
    if type(key) is int:
        key = [key]
    for n,k in enumerate(key):
        for y in range(1, ws.max_row+1):
            cell = ws.cell(column=k, row=y)
            cell.fill = color[n]
    return



def sort_data2( ws, keys):
    """
    ワークシートのデータをソートする関数

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): ワークシートオブジェクト
        keys: (1,2) col number (1より開始) A:1, B:2 .... (1,2)
    return 
        ws: ソートされたもの
    """
    print("D0 keys={keys}")
    
    data = list(ws.iter_rows(min_row=2))  # ヘッダーを除いたデータを取得
    print(f"data={data}")
    for i, row in enumerate(data, start=2):
        for j, cell in enumerate(row):
            if j==0: print(f"{cell}, value={cell.value}")
    k1, k2 = keys
    data.sort(key=lambda row: (row[k1-1].value, row[k2-1].value))  # B列と10列目でソート
    print(f"D1 data={data}")
    
    # ソートされたデータをワークシートに書き込む
    for i, row in enumerate(data, start=2):
        for j, cell in enumerate(row):
            if j==0: print(f"{cell}, value={cell.value}")
            ws.cell(row=i, column=j + 1, value=cell.value)
    return ws

if __name__ == "__main1__":
    wb, ws, ws2 = copy_worksheet_with_styles()
    ws2 = sort_data2(ws2, keys=(1, 3))
    wb.save("mylib\\bbb.xlsx")
    sys.exit()
    
def excel_copy_sort_search(input_file="mylib\\aaa.xlsx",
                           output_file="mylib\\bbb.xlsx", search_val="1234"):
    """
    Excelファイルのコピー、ソート、検索を行う関数

    Args:
        input_file (str): 入力Excelファイル名
        output_file (str): 出力Excelファイル名
        search_val (str): 検索する値
    """
    #if 1:
    try:
        # Excelファイルの読み込み
        wb = openpyxl.load_workbook(input_file)
        ws = wb["test1"]
        print("C1")
        # 新しいワークシートを作成し、データをコピー
        ws2 = wb.create_sheet(title="test2")
        for row in ws.iter_rows(values_only=True):
            ws2.append(row)
        print("C2")
        # データのソート
        sort_data(ws2)
        print("C3")
        # ソート結果を保存
        wb.save(output_file)

        # 検索
        col_no = 1  #1:A  2:B列
        n = 2  # 2行目から検索開始
        col, line = search(ws2, col_no, search_val, n)

        if col:
            print(f"値 '{search_val}' が行 {line} に見つかりました。")
            print(f"列データ: {col}")
        else:
            print(f"値 '{search_val}' は見つかりませんでした。")

    except FileNotFoundError:
        print(f"ファイル '{input_file}' が見つかりません。")
    except KeyError:
        print("指定されたシートが見つかりません。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")


def sort_data(ws):
    """
    ワークシートのデータをソートする関数

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): ワークシートオブジェクト
    """
    print("D0")
    data = list(ws.iter_rows(min_row=2))  # ヘッダーを除いたデータを取得
    data.sort(key=lambda row: (row[0].value, row[3].value))  # B列と10列目でソート
    print("D1")
    # ソートされたデータをワークシートに書き込む
    for i, row in enumerate(data, start=2):
        for j, cell in enumerate(row):
            ws.cell(row=i, column=j + 1, value=cell.value)


def search(ws, col_no, val, n):
    """
    ワークシート内で指定された値を検索する関数

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): ワークシートオブジェクト
        col_no (int): 検索する列番号
        val (str): 検索する値
        n (int): 検索を開始する行番号

    Returns:
        tuple: 見つかった場合は列データのリストと行番号、見つからなかった場合はNone, None
    """
    print(f"#search: #151: {val}")
    for row in ws.iter_rows(min_row=n):
        if str(row[col_no - 1].value) == val:
            col = [cell.value for cell in row]
            return col, row[0].row
    return None, None


if __name__ == "__main1__":
    import os, sys
    script_path = os.path.abspath(sys.argv[0])
    script_name = os.path.basename(script_path)
    current_directory = os.getcwd()
    print("現在のディレクトリ:", current_directory)
    print(f"=== Start: {script_name} ===")
    print(f"スクリプトのパス: {script_path}")
    print(f"スクリプト名: {script_name}")

    excel_copy_sort_search(search_val="3")

#
# ======================================================================================
#  Excel ファイルの読み込み

def openXl(fn, sheet_name="JROD"):
    """
    # Excel ファイルを開く
    Args:
        fn: file name
        sheet_name: sheet name

    Returns:
        workbook,
        worksheet,
        title: ヘッダーのカラム名 {1:'title1', 2:'title2, ... 'max_row':21, 'max_column':114}

    """
    workbook = load_workbook(fn)
    sheet = workbook[sheet_name]  # シート名を指定
    # info
    print(f"<openXl> fn: {fn}, sheet: {sheet_name}, max row={sheet.max_row}, col={sheet.max_column}")
    
    # セル (x, y) の値を取得 (x, y は 1 始まり)
    if 0:
        x = 2  # 例: 2 行目
        y = 3  # 例: 3 列目
        for y in range(1,10):
            print(f'{sheet.cell(row=1, column=y).value:16}:', end="")
            cell_value = sheet.cell(row=x, column=y).value
            print(f"セル({x}, {y}) の値: {cell_value}")

    # title set
    title = {}
    for y in range(1, sheet.max_column+1):
        title[y] = sheet.cell(row=1, column=y).value
    print(f"<openXl> title 1:{title[1]}, title max:{title[sheet.max_column]}")
    title['max_row'] = sheet.max_row
    title['max_column'] = sheet.max_column

    return workbook, sheet, title

    # すべてのセルを読み込む (ジェネレータ)
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value)

# ==============================================================
# Excel ファイルへの書き込み
'''
def writeXl():
    # Excel ファイルを開く
    workbook = load_workbook('your_excel_file.xlsx')

    # シートを取得
    sheet = workbook['Sheet1']

    # セル (x, y) に値を書き込む (x, y は 1 始まり)
    x = 2  # 例: 2 行目
    y = 3  # 例: 3 列目
    sheet.cell(row=x, column=y).value = '新しい値'

    # Excel ファイルを保存
    workbook.save('your_excel_file.xlsx')
'''
# ================================================================
# 特定のセル範囲の読み書き
'''
def readWrite():

    # Excel ファイルを開く
    workbook = load_workbook('your_excel_file.xlsx')

    # シートを取得
    sheet = workbook['Sheet1']

    # セル範囲 (x1, y1) から (x2, y2) の値を読み込む
    x1 = 2
    y1 = 3
    x2 = 4
    y2 = 5
    for row in sheet.iter_rows(min_row=x1, min_col=y1, max_row=x2, max_col=y2):
        for cell in row:
            print(cell.value)

    # セル範囲 (x1, y1) から (x2, y2) に値を書き込む
    x1 = 2
    y1 = 3
    x2 = 4
    y2 = 5
    values = [['a', 'b', 'c'], ['d', 'e', 'f']]  # 書き込む値のリスト
    for i, row in enumerate(sheet.iter_rows(min_row=x1, min_col=y1, max_row=x2, max_col=y2)):
        for j, cell in enumerate(row):
            cell.value = values[i][j]

    # Excel ファイルを保存
    workbook.save('your_excel_file.xlsx')
'''
# ================
# wsの指定した行のレコードを取得する。
def getRow(ws, x):
    """
    ワークシートの指定した行のすべての列の値を取得する関数

    Args:
        ws: openpyxlのワークシートオブジェクト
        x: 取得する行番号 (1始まり)

    Returns:
        指定した行のすべての列の値を含むリスト. d[0] = row_no
        行が存在しない場合は空のリストを返す
    """
    if x > ws.max_row: x = ws.max_row
    row_values = [x]
    if 1 <= x <= ws.max_row:  # 行が存在するか確認
        for cell in ws[x]:  # 指定した行のすべてのセルを反復処理
            if cell.value is None:
                v = ""
            else: v = cell.value
            row_values.append(v)  # セルの値をリストに追加
    return row_values

def setRow(ws, x, data):
    """
    1行分のデータを指定した行のワークシートに書き込む関数

    Args:
        ws: openpyxlのワークシートオブジェクト
        x: 書き込む行番号 (1始まり)
        data: 書き込むデータを含むリスト
    """
    if data[0] != x:
        print(f"ERROR: list[0] mismatch x:{x}, list[0]={data[0]}")
        return
    if 1 <= x <= ws.max_row + 1:  # 行が存在するか、または次の行に書き込むか確認
        for y, value in enumerate(data):  # データを列方向に書き込む
            if y == 0: continue
            ws.cell(row=x, column=y).value = value  # セルに値を書き込む
    else:
        print(f"Error: Row {x} is beyond the worksheet boundaries.")  # エラーメッセージを表示
    


if __name__ == "__main1__":
    print("このスクリプトは直接実行されました")
    fn = 'JRODe_SRC_Sample.xlsx'
    print(f'excel filename: {fn}')
    wb, ws, title = openXl(fn)
    #print(title)
    
    rows = getRow(ws, 1)
    print("\nROW:1, データを表示(y:1 ... 10)")
    for i in range(1, 10):
        print(f'{i}, {title[i]:10}: {rows[i]}')
        
    #print("rows=", rows)    

def logHeaderSet(ws_log):
    # --- color and width
    ws_log.append(log_header)
    
    for col in range(1, ws_log.max_column+1):
        xlwidth(ws_log, col, log_width[col-1])
        xlcolor(ws_log, 1, col, )
    # ---


def createLogWs(wb, ws):
    """
    wbに "copy" sheetや "ws_log"を作成。
    既に、あれば、それぞれに、ws2, ws3を紐付ける

    Args:
        wb: 既に開いてある、ワークブックプロジェクト
        ws: worksheet
    return:
        w2, w3: 紐付けられた、ワークシートオブジェクト
    """
    # sheet名をチェック。
    sheet_names = [ws.title for ws in wb.worksheets]
    ws_title_list = sorted([ws.title for ws in wb.worksheets])
    ws_length = len(ws_title_list) - 1
    print(f"line() sheet_names = {sheet_names},\nwb.sheetnames = {wb.sheetnames}\n ws_length={ws_length}")
    # ws2
    ws2_sheet = None
    for s in wb.sheetnames:
        if "_copy" in s:
            print("found: sheet name=", s)
            ws2_sheet = s
            break
    if ws2_sheet is not None:
        print(f"{line()} ws2('{ws2_sheet}') sheet opened.")
        ws2 = wb[ws2_sheet]
    else:
        print(f"{line()} 'ws_copy' created.")
        #ws2 = wb.create_sheet("ws_copy")
        ws2 = wb.copy_worksheet(ws)
    # ws3
    if "ws_log" in sheet_names:
        print(f"{line()} 'ws_log' opened.")
        ws3 = wb["ws_log"] # シートの取得
    else:
        print(f"{line()} 'ws_log' created.")
        ws3 = wb.create_sheet(title="ws_log") # 追加するシート名    
        logHeaderSet(ws3)
    return ws2, ws3


import inspect
import os

def location():
    frame = inspect.currentframe().f_back
    return os.path.basename(frame.f_code.co_filename), frame.f_code.co_name, frame.f_lineno

def line():
    frame = inspect.currentframe().f_back
    return f"{frame.f_code.co_name}: #{frame.f_lineno:3}:"

if __name__ == "__main__":
    print("このスクリプトは直接実行されました\n")
    
    print(location())
    
    fn = 'JRODe_SRC_Sample.xlsx'
    fn = "JRODe_test.xlsx"
    fn = "JRODe_ARC_Sample.xlsx"
    
    # book open and info
    print(f'excel filename: {fn}')
    wb, ws, title = openXl(fn, "Sheet1")
    #print(f"{line()} title = {title}")
    print(f"{line()} {fn} was opend.\n  dimensions= {ws.dimensions}")
    ts = list(title.items())
    #print(f"{line()} ts= {ts}")
    print(f"#551 title = {ts[:10]} ...")
    
    ws2, ws3 = createLogWs(wb, ws)
    
    # search
    if 0:
        col, n = search(ws, 1, "2023-0617", 2)
        if col is not None:
            print(f"#558 found: {col[:10]}, {n}")
        else:
            print(line()," not found.")
    '''
    # ws create
    ws2 = wb.copy_worksheet(ws)
    ws_log = wb.create_sheet(title="ws_log")
    # --- color and width
    logHeaderSet(ws_log)
    '''
    # --- trans
    if 0:
        trans(ws2, 3, 14, "test-dat", ws_log, "test")
    
    
    # --- save
    fn_color = "mylib\\color_and_width.xlsx"
    wb.save(fn_color)
    
    
    print(f"Normal end. Saved to '{fn_color}'.")