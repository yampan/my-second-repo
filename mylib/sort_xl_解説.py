### sort_xl.py

#### Excel作成/編集を自動化！python openpyxlのまとめ
# https://qiita.com/mathlive/items/20078f4b31273c180f51
# 公式ドキュメントは以下です。
# https://openpyxl.readthedocs.io/en/stable/
# 
'''
Excelファイル新規作成
wb = openpyxl.Workbook() # workbookの作成

既存Excelファイル読み込み

wb = openpyxl.load_workbook("test.xlsx") # Excelファイルの読み込み

wb.save("out.xlsx") # Excelファイルの保存

## Worksheet操作

    wb.create_sheet(title="追加するシート名")
    wb.create_sheet(title="追加するシート名", index=追加位置)

sheet_names = wb.sheetnames # シート名一覧取得

ws = wb["test1"] # シートの取得
ws.title = "テスト名変更" # 新しいシート名
wb.move_sheet("test1", offset=1) # シートの移動

ws = wb.copy_worksheet(wb["test2"]) # ws copy
# コピー後のシート名はコピー元のシート名の末尾に「Copy」がつきます。


min_row = ws.min_row # 行の最小
max_row = ws.max_row # 行の最大
min_col = ws.min_column # 列の最小
max_col = ws.max_column # 列の最大

### cell access

cell = ws["アルファベットと数字の組み合わせ"] # A1
cell = ws.cell(row=1, column=1) # A1セルの取得

 列のアルファベット名と数値の変換
column_letter = openpyxl.utils.get_column_letter(10)
column = openpyxl.utils.column_index_from_string("D")

row = ws[行番号] # 1
for cell in row:
    cell の処理

for row in ws: # 各行のループ
    for cell in row:
        cell = ...

col = ws["A"] # 1列取得
for cell in col:
    cellごとの処理

for col in ws.columns # 各列でループ
    for cell in col:
        cellごとの処理

subset_ws = ws.iter_rows(min_row, max_row, min_col, max_col)
subset_ws = ws["アルファベットと数字の組み合わせ:アルファベットと数字の組み合わせ"]

for row in subset_ws: # subset_ws内の各行ごとにループ
    for cell in row:
        cellごとの処理


cell.value

セルから座標を取得することもできます。
cell = ws["A1"] # A1セルの取得
print(f"セルアドレス : {cell.coordinate}") # A1
print(f"セル列番号 : {cell.column}")       # 1
print(f"セル列アルファベット : {cell.column_letter}") # A
print(f"セル行番号 : {cell.row}")  # 1

# セルのスタイル変更
cell = ws["B2"]
border = cell.border # 罫線
fill = cell.fill # 背景色
font = cell.font # 文字

openpyxl.styles.Side(style="線のスタイル", color="色コード")
openpyxl.styles.Border(top=上の線, bottom=下の線, left=左の線, right=右の線)

line = openpyxl.styles.Side(style="thick", color="000000") # 太線・黒色
border = openpyxl.styles.Border(top=line, bottom=line, left=line, right=line) # lineを上下左右すべてに適用

cell = ws["F4"]
cell.border = border
wb.save("output.xlsx")

### 背景色の変更
openpyxl.styles.PatternFill(fgColor="前面の色", bgColor="背面の色", fill_type="塗り方")
fill_color = openpyxl.styles.PatternFill(fgColor="B8CCE4", bgColor="B8CCE4", fill_type="solid") # 青色単色
cell = ws["F4"]
cell.fill = fill_color

### font
font = openpyxl.styles.fonts.Font(color="FF0000", size=20) # 赤色 size=20
                                  name="HGP創英角ﾎﾟｯﾌﾟ体"
                                  bold=True, italic=True
cell = ws["B2"]
cell.font = font

cell = ws["B5"]
cell.number_format = "0.000" # 0.000表記
                      #,##0.00 	1,234.57
                      00000 	01235
                      hh:mm:ss  22:33:44	
                      "yyyy-mm-dd (aaa)"   2025-02-22 (土)
### cell 幅
ws.column_dimensions["B"].width = 10

### 行挿入
ws.insert_rows(1)
ws.delete_rows(1)  #削除

ws.insert_cols(1)
ws.delete_cols(1)

 セル幅自動調整
 https://qiita.com/mathlive/items/20078f4b31273c180f51#5-%E3%81%8A%E3%81%BE%E3%81%91%E3%82%B9%E3%83%8B%E3%83%9A%E3%83%83%E3%83%88


## シートの移動
move_sheet = "移動したいシート名"
offset = -wb.sheetnames.index(move_sheet) # 先頭までの個数を取得
wb.move_sheet(move_sheet, offset=offset) # 先頭へのシートの移動

move_sheet = "移動したいシート名"
offset = len(wb.sheetnames)-wb.sheetnames.index(move_sheet)-1 # 最後尾までの個数を取得
wb.move_sheet(move_sheet, offset=offset) # 最後尾へのシートの移動
'''



### Excelシートを昇順で並べる
"""
Excelシートを昇順で並べる


# Excelファイルの読み込み
wb = load_workbook(selected_file)

# Excelシート一覧（昇順）
ws_title_list = sorted([ws.title for ws in wb.worksheets])
ws_length = len(ws_title_list) - 1

# Excelシート並び替え実行
for ws_title in ws_title_list:
    ws = wb[ws_title]

    for row in range(ws.max_row):
        if ws.max_row != 1:
            for col in range(ws.max_column):
                ws.cell(row=row+1, column=col+1).font = font

    wb.move_sheet(ws, offset=ws_length)

wb.save(selected_file.replace(file_name, f"sorted_{file_name}"))
"""


"""
# ワークブックを開く
wb = load_workbook('sample.xlsx')

# アクティブなシートを取得
ws = wb.active

# 1行目のすべてのセルにアクセス
for cell in ws[1]:
    print(cell.value)

# 'A'列のすべてのセルにアクセス
for cell in ws['A']:
    print(cell.value)
"""


#### セルの書式設定
"""
# A1セルに値を書き込む
ws['A1'] = 'Hello, Excel!'

# A1セルのフォントを変更
ws['A1'].font = Font(name='Arial', size=20, bold=True, color=Color(rgb="FFFFFF00"))

# A1セルの背景色を設定
ws['A1'].fill = PatternFill(fill_type="solid", fgColor="0000FF00")
"""







from openpyxl import Workbook, styles
def sort_test1():
    
    # フィルタを適用
    ws.auto_filter.ref = "A1:C5"

    # ソートの条件を設定（例：Price列を昇順にソート）
    ws.auto_filter.add_sort_condition("B2:B5")

    # ファイルを保存
    wb.save('mylib\\sort_test.xlsx')

def row2list(row, deb=0):
    l = []
    if deb: print("#214: row=", row)
    for cell in row:
        if deb: print("#216: cell=", cell.value)
        l.append(cell.value)
    return l

### ソート
def sort2(wb, ws, key, deb=0):
    """
    sort keyは1 or 2個。row dataの最後にindexを付加
    args:
        wb: workbook
        ws: worksheet
        key:int or list -  key no (1始まり)
    return:
        ws: sorted ws
    """
    # get all cell
    header = []
    data = []
    index_col = ws.max_column+1
    for n, row in enumerate(ws, start=1): # 各行のループ
        cell = ws.cell(row=n, column=index_col)
        if n == 1: 
            header = row
            cell.value = 'index'
            header = header + (cell,)
        else:
            cell.value=n
            row = row + (cell,)
            data.append(row)
    # ---
    if deb:
        print(f"HDR {header}")
        for n, i in enumerate(data, start=2):
            print(f"{n}): {i}")
    
    # sort
    if type(key) is int:
        data.sort(key=lambda row: (row[key-1].value)) 
    else:
        data.sort(key=lambda row: (row[key[0]-1].value,  row[key[1]-1].value)) 
    print(f"#252 key = {key}")
    for n, i in enumerate(data, start=2):
        print(f"{n}): {i}")
           
    #ws2['A1'] = row2list(header)[0] # OK
    #ws2[1] = row2list(header) # NG
    #print(f"#260 'A1:A2' = {ws2['A1:B1']}") # Cell get OK, cell write NG
    
    # 各セルから値を抽出して、　data_v に保存する。
    data_v = []
    a = []
    for i in header:
        a.append(i.value) 
    data_v.append(a)
    for row in data:
        a = []
        for col in row:
            a.append(col.value)
        data_v.append(a)
        
    # data_v から値をもどす。
    for y, row in enumerate(data_v, start=1):
        for x, col in enumerate(row, start=1):
            ws.cell(row=y, column=x).value = col
    
    # sort column に色を付ける
    fill_color1 = styles.PatternFill(fgColor="ACEBF0", fill_type="solid") # 青色単色
    fill_color2 = styles.PatternFill(fgColor="FFFF66", fill_type="solid") # 黄色
    color = [fill_color1, fill_color2]
    if type(key) is int:
        key = [key]
    for n,k in enumerate(key):
        for y in range(1, ws.max_row+1):
            cell = ws.cell(column=k, row=y)
            cell.fill = color[n]
    return

if __name__ == "__main__":
    # 新しいワークブックを作成
    wb = Workbook()
    ws = wb.active

    # データを挿入
    data = [
        ["Item", "Price", "Quantity"],
        ["Apple", 0.5, 10],
        ["Banana", 0.25, 20],
        ["Cherry", 1.0, 15],
        ["Banana", 0.5, 19],
        ["Date", 1.5, 5]
    ]
    for row in data:
        ws.append(row)
    
    # sheet copy
    ws2 = wb.copy_worksheet(ws)
    ws2.title = 'Sorted'
    
    key = [1,3] # sort-key (1始まり)
    sort2(wb, ws2, key)
    
    # ファイルを保存
    fn_save = 'mylib\\sort_test.xlsx'
    wb.save(fn_save)
    print(f"#255 saved to '{fn_save}'. (key={key})")
    # ---
    print("Normal end.")